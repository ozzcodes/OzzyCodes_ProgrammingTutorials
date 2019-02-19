from alpha_vantage.timeseries import TimeSeries
import openpyxl
import datetime

# Get json object with the intraday data and another with  the call's metadata
# data, meta_data = ts.get_intraday(symbol='MSFT',interval='1day', outputsize='full')
# data, meta_data = ts.get_intraday('GOOGL', interval='day')
# data.describe()
CST_ExcelName = "stockData.xlsx"
CST_ExcelSheetName = "TSLA"
tupleXLSX = {}


class stock:
    symbol = None
    close = 0
    open = 0
    low = 0
    high = 0
    max = 0
    sl = 0
    updated = 0
    support = 0
    fsupport = 0

    def __init__(self):
        self.symbol = None
        self.close = 0
        self.open = 0
        self.low = 0
        self.high = 0
        self.max = 0
        self.sl = 0
        self.updated = 0
        self.support = 0
        self.fsupport = 0


class importXLSX:
    _symbols = []
    _xlsx = None
    _sheet = None
    _ExcelName = None

    def __init__(self, ExcelName=None, ExcelSheetName=None):
        if ExcelName is not None:
            self.openXLS(ExcelName, ExcelSheetName)
            self._ExcelName = ExcelName

    def __loadXLSX(self):
        try:
            sym = self._sheet.cell(row=2, column=1)
            for r in range(2, self._sheet.max_row + 1):
                key = self._sheet.cell(row=r, column=1).value
                strd = self._sheet.cell(row=r, column=3).value
                if strd is None: strd = "1900-01-01"
                date = datetime.datetime.strptime(strd, "%Y-%m-%d")
                if date.date() < datetime.datetime.now().date():
                    self._symbols.append(key)
                else:
                    self._symbols.append('$NONE$')
                # _symbols[locals()["key"]] = 0
            # print (self._symbols)
        except FileNotFoundError:
            print("Unable to load XLSX.")

    def update_symbol(self, my_symbol, values=stock):
        if my_symbol == '$NONE$':
            return

        idx = self._symbols.index(my_symbol) + 2
        self._sheet.cell(row=idx, column=3, value=values.updated)
        self._sheet.cell(row=idx, column=4, value=values.open)
        self._sheet.cell(row=idx, column=5, value=values.high)
        self._sheet.cell(row=idx, column=6, value=values.close)
        self._sheet.cell(row=idx, column=7, value=values.low)
        self._sheet.cell(row=idx, column=8, value=values.max)
        self._sheet.cell(row=idx, column=9, value=values.support)
        self._sheet.cell(row=idx, column=10, value=values.fsupport)

    def get_symbol(self):
        return self._symbols

    def openXLS(self, ExcelName, ExcelSheetName):
        self._ExcelName = ExcelName
        self._xlsx = openpyxl.load_workbook(ExcelName)
        self._sheet = self._xlsx.get_sheet_by_name(ExcelSheetName)
        self.__loadXLSX()

    def closeXLSX(self):
        self._xlsx.close()

    def saveXLSX(self):
        try:
            self._xlsx.save(self._ExcelName)
        except ValueError:
            print("Unable to extract data to XLSX")


class finance_series:
    ts = TimeSeries(key='JGQU9ILE9JTG8CAG', output_format='pandas')
    idx = 0
    data = None
    meta_data = None

    def __init__(self, msymbol=""):
        self.idx = 0
        if msymbol != "":
            self.LoadDataFromSymbol(msymbol)

    def LoadDataFromSymbol(self, symbolToLoad, interval='daily'):
        print("Getting data for %s..." % symbolToLoad)
        try:
            if interval == 'daily':
                self.data, self.meta_data = self.ts.get_daily_adjusted(symbol=symbolToLoad, outputsize='compact')
            elif interval == 'weekly':
                self.data, self.meta_data = self.ts.get_weekly_adjusted(symbol=symbolToLoad)
        except ValueError:
            print("Error printing the correct stock symbol %s" % symbolToLoad)

    def setIndex(self, v_idx):
        self.idx = v_idx

    def getData(self):
        return self.data

    def getMetaData(self):
        return self.meta_data

    def maxHighHistoric(self):
        r = self.data.loc[self.data['2. high'] == self.data.max()['2. high']]
        return [r.index[0], r.iloc[-1]['2. high']]

    def high(self, _idx=None):
        if _idx == "": _idx = self.idx
        return self.data.iloc[_idx]['2. high']

    def close(self, _idx=None):
        if _idx == "": _idx = self.idx
        return self.data.iloc[_idx]['4. close']

    def low(self, _idx=None):
        if _idx is None: _idx = self.idx
        return self.data.iloc[_idx]['3. low']

    def open(self, _idx=None):
        if _idx is None: _idx = self.idx
        return self.data.iloc[_idx]['1. open']

    def date(self, _idx=None):
        if _idx is None: _idx = self.idx
        return self.data.index[_idx]

    def locateSupport(self):
        support = self.low(0)
        resistance = self.high(0)
        new_max = resistance
        new_min = support
        # data.iloc[1:5]['2. high'].max()
        ind = 0
        for d in range(self.data.count()[1]):
            l = self.low(d)
            h = self.high(d)
            if h > new_max:
                new_max = h
            if l < new_min:
                new_min = l
                resistance = new_max
            if h > resistance:
                support = new_min
                new_min = l
                ind = d
            # print ("[%i] S= %f R= %f l= %f h = %f " %(d, support, resistance, l, h))
            # if d > 40: break
        return [self.data.index[ind], support]


def main():
    print("Stock Portfolio")
    fs = finance_series()
    my_xlsx = importXLSX()

    my_xlsx.openXLS(CST_ExcelName, CST_ExcelSheetName)
    my_symbol = my_xlsx.get_symbol()
    for s in my_symbol:
        if s == '$NONE$':
            continue

        ticker = stock
        fs.LoadDataFromSymbol(s, 'weekly')
        ticker.symbol = s
        ticker.open = fs.open(-1)
        ticker.close = fs.close(-1)
        ticker.low = fs.low(-1)
        ticker.high = fs.high(-1)
        ticker.updated = fs.date(-1)
        ticker.max = fs.maxHighHistoric()[1]
        support = fs.locateSupport()
        ticker.support = support[1]
        ticker.fsupport = support[0]

        my_xlsx.update_symbol(s, ticker)

        print("Support is %s value %f in %s" % (s, support[1], support[0]))
        # break
    my_xlsx.saveXLSX()
    my_xlsx.closeXLSX()
    exit(0)

    fs.LoadDataFromSymbol('WB', 'weekly')
    mh = fs.maxHighHistoric()
    meta_data = fs.getMetaData()

    print("max: Date=%s, value=%f" % (str(mh[0]), mh[1]))
    for i in range(10, 20):
        fs.setIndex(i)
        print(str(fs.open()))

    # pprint(meta_data)

    exit(0)

    # pprint(lows.tail(5))
    # pprint(data.describe())
    # pandas.iloc[file, col]
    # data.columns = ['open', 'high', 'low', 'close', 'aclose', 'volume']
    # del data['aclose']
